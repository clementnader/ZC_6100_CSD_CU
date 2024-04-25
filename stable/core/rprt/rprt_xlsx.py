# -*-coding:utf-8 -*
import sys
import os
from datetime import datetime

# Openpyxl lib imports
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet import dimensions
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.prod.sw import SwConst, SwConstResult


class ReporterXlsx(object):

    str_dir_path = None
    str_sys_sw_sw_ver = None
    work_book = None

    def __init__(self, str_dir_path, str_sys_sw_sw_ver):
        self.str_dir_path = str_dir_path
        self.str_sys_sw_sw_ver = str_sys_sw_sw_ver
        self.work_book = Workbook()

    def report_constr(self, d_sw_constr, d_sw_constr_result):

        for str_constr_id in d_sw_constr:

            str_constr_type = d_sw_constr[str_constr_id]['type']

            print(f'Reporting constraint: {str_constr_id}, type: {str_constr_type}')

            str_ws_title = str_constr_id.replace('/', '_')
            work_sheet = self.work_book.create_sheet(str_ws_title)

            self._set_page_format(work_sheet, str_constr_type)

            self._add_constraint_id(work_sheet, str_constr_id)
            str_constr_type_text = SwConst.get_text(str_constr_type)
            self._add_constraint_type(work_sheet, str_constr_type_text)

            i_idx_row = self._add_verif_banner(work_sheet, str_constr_type)

            i_idx_row = self._report_verif_constr(work_sheet, i_idx_row, str_constr_type, d_sw_constr[str_constr_id])

            e_result = d_sw_constr_result[str_constr_id]['result']

            i_idx_row = self._add_result_banner(work_sheet, i_idx_row, e_result, str_constr_type)

            self._report_result_constr(work_sheet, i_idx_row, str_constr_type, d_sw_constr_result[str_constr_id])

        self._save()

    def _report_verif_constr(self, work_sheet, i_idx_row, str_constr_type, d_sw_constr):

        if SwConst.SW_CONST_EXE_MAP_LIM_SECT == str_constr_type:
            i_idx_row = self._report_verif_constr_map_lim_sect(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_EXE_MAP_OFF_SECT == str_constr_type:
            i_idx_row = self._report_verif_constr_map_off_sect(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_EXE_MAP_OCCUR_SYMB == str_constr_type:
            i_idx_row = self._report_verif_constr_map_occur_symb(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_EXE_INSTR_MAND == str_constr_type:
            i_idx_row = self._report_verif_constr_instr_mand(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_EXE_INSTR_FORB == str_constr_type or SwConst.SW_CONST_EXE_INSTR_OCCUR == str_constr_type:
            i_idx_row = self._report_verif_constr_instr_forb_occur(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_EXE_INSTR_OFF == str_constr_type:
            i_idx_row = self._report_verif_constr_instr_off(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_EXE_COMM_STR == str_constr_type:
            i_idx_row = self._report_verif_constr_comm_str(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_LOG_COMP_OPT == str_constr_type:
            i_idx_row = self._report_verif_constr_log_comp_opt(work_sheet, i_idx_row, d_sw_constr)
        elif SwConst.SW_CONST_LOG_OCCUR_WARN == str_constr_type:
            i_idx_row = self._report_verif_constr_log_occur_warn(work_sheet, i_idx_row, d_sw_constr)
        else:
            pass

        return i_idx_row

    def _report_verif_constr_map_lim_sect(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1

        for str_sect_name, d_lim_sect in d_sw_constr['sect'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = str_sect_name + ' section'
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Start symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Start offset symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':C' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('E' + str_idx_row + ':F' + str_idx_row)

            curr_cell = work_sheet['H' + str_idx_row]
            curr_cell.fill = cell_fill

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Section'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['C' + str_idx_row]
            curr_cell.value = 'Symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Section'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['F' + str_idx_row]
            curr_cell.value = 'Symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['H' + str_idx_row]
            curr_cell.value = 'Incl./Excl. address'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Consolas', size='11')

            if 'start' in d_lim_sect:
                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = d_lim_sect['start']['symb_sect']
                curr_cell.font = cell_font

                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = d_lim_sect['start']['symb_name']
                curr_cell.font = cell_font

                if 'symb_off_sect' in d_lim_sect['start'] and 'symb_off_name' in d_lim_sect['start']:
                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_lim_sect['start']['symb_off_sect']
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['F' + str_idx_row]
                    curr_cell.value = d_lim_sect['start']['symb_off_name']
                    curr_cell.font = cell_font
                else:
                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['F' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font

                curr_cell = work_sheet['H' + str_idx_row]

                if d_lim_sect['start']['symb_addr_incl']:
                    curr_cell.value = 'Included address'
                else:
                    curr_cell.value = 'Excluded address'

                curr_cell.font = Font(name='Calibri', size='11')

            else:
                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['F' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['H' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'End symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'End offset symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':C' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('E' + str_idx_row + ':F' + str_idx_row)

            curr_cell = work_sheet['H' + str_idx_row]
            curr_cell.fill = cell_fill

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Section'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['C' + str_idx_row]
            curr_cell.value = 'Symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Section'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['F' + str_idx_row]
            curr_cell.value = 'Symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['H' + str_idx_row]
            curr_cell.value = 'Incl./Excl. address'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Consolas', size='11')

            if 'end' in d_lim_sect:
                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = d_lim_sect['end']['symb_sect']
                curr_cell.font = cell_font

                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = d_lim_sect['end']['symb_name']
                curr_cell.font = cell_font

                if 'symb_off_sect' in d_lim_sect['end'] and 'symb_off_name' in d_lim_sect['end']:
                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_lim_sect['end']['symb_off_sect']
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['F' + str_idx_row]
                    curr_cell.value = d_lim_sect['end']['symb_off_name']
                    curr_cell.font = cell_font
                else:
                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font
                    curr_cell = work_sheet['F' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font

                curr_cell = work_sheet['H' + str_idx_row]

                if d_lim_sect['end']['symb_addr_incl']:
                    curr_cell.value = 'Included address'
                else:
                    curr_cell.value = 'Excluded address'

                curr_cell.font = Font(name='Calibri', size='11')

            else:
                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['F' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font
                curr_cell = work_sheet['H' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Excluded symbols'
            curr_cell.font = Font(name='Calibri', size='11', bold=True)
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':C' + str_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]

            if 'excl' in d_lim_sect:
                curr_cell.value = ', '.join(d_lim_sect['excl'])
            else:
                curr_cell.value = '-'

            curr_cell.font = Font(name='Consolas', size='11')

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 7

            i_idx_row += 1

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_map_off_sect(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1

        for str_sect_name, d_off_sect in d_sw_constr['sect'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = str_sect_name + ' section'
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Files / Address offsets'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Reference file'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':C' + str_idx_row)

            curr_cell = work_sheet['D' + str_idx_row]
            curr_cell.value = 'Other file'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Address offset'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = d_off_sect['ref']
            curr_cell.font = Font(name='Calibri', size='11')

            for str_file_name, str_addr_off in d_off_sect['off'].items():

                curr_cell = work_sheet['D' + str_idx_row]
                curr_cell.value = str_file_name
                curr_cell.font = Font(name='Calibri', size='11')

                curr_cell = work_sheet['E' + str_idx_row]
                i_addr_off = int(str_addr_off, 16)
                curr_cell.value = "0x%08x" % i_addr_off
                curr_cell.font = Font(name='Consolas', size='11')

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Excluded symbols'
            curr_cell.font = Font(name='Calibri', size='11', bold=True)
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]

            if 'excl' in d_off_sect:
                curr_cell.value = ', '.join(d_off_sect['excl'])
            else:
                curr_cell.value = '-'

            curr_cell.font = Font(name='Consolas', size='11')

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 7

            i_idx_row += 1

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_map_occur_symb(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell_font = Font(name='Calibri', size='11', bold=True)

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Section'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

        curr_cell = work_sheet['C' + str_idx_row]
        curr_cell.value = 'Symbol'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('C' + str_idx_row + ':E' + str_idx_row)

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_font = Font(name='Consolas', size='11')

        for str_sect_name, l_symbols in d_sw_constr['symb'].items():

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = str_sect_name
            curr_cell.font = cell_font

            for str_symbol in l_symbols:

                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = str_symbol
                curr_cell.font = cell_font

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_instr_mand(self, work_sheet, i_idx_row, d_sw_constr):

        for str_file_name, d_symb_name_addr_off_const_instr in d_sw_constr['instr'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Symbols'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['G' + str_idx_row]
            curr_cell.value = 'Instructions'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('G' + str_idx_row + ':H' + str_idx_row)

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

            curr_cell = work_sheet['G' + str_idx_row]
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('G' + str_idx_row + ':J' + str_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Name'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

            curr_cell = work_sheet['G' + str_idx_row]
            curr_cell.value = 'Offset address'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['H' + str_idx_row]
            curr_cell.value = 'Mnemonic'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['I' + str_idx_row]
            curr_cell.value = 'Field name'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['J' + str_idx_row]
            curr_cell.value = 'Field value'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            for str_symb_name in d_symb_name_addr_off_const_instr:
                curr_cell = work_sheet['A' + str_idx_row]

                cell_font = Font(name='Consolas', size='11')

                curr_cell.value = str_symb_name
                curr_cell.font = cell_font

                for str_instr_addr_off, d_const_instr in d_symb_name_addr_off_const_instr[str_symb_name].items():

                    curr_cell = work_sheet['G' + str_idx_row]
                    i_instr_addr_off = int(str_instr_addr_off, 16)
                    curr_cell.value = "0x%08x" % i_instr_addr_off
                    curr_cell.font = cell_font

                    cell_font = Font(name='Calibri', size='11')

                    curr_cell = work_sheet['H' + str_idx_row]
                    curr_cell.value = d_const_instr['mnem']
                    curr_cell.font = cell_font

                    if 'fields' in d_const_instr:
                        for str_field_name in d_const_instr['fields']:
                            curr_cell = work_sheet['I' + str_idx_row]
                            curr_cell.value = str_field_name
                            curr_cell.font = cell_font

                            curr_cell = work_sheet['J' + str_idx_row]
                            curr_cell.value = d_const_instr['fields'][str_field_name]
                            curr_cell.font = cell_font

                            i_idx_row += 1
                            str_idx_row = str(i_idx_row)
                    else:
                        curr_cell = work_sheet['I' + str_idx_row]
                        curr_cell.value = '-'
                        curr_cell.font = cell_font

                        curr_cell = work_sheet['J' + str_idx_row]
                        curr_cell.value = '-'
                        curr_cell.font = cell_font
                        i_idx_row += 1
                        str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7
            i_idx_row += 1

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_instr_forb_occur(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_font = Font(name='Calibri', size='11', bold=True)
        cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Mnemonic'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':b' + str_idx_row)

        curr_cell = work_sheet['C' + str_idx_row]
        curr_cell.value = 'Field name'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        curr_cell = work_sheet['D' + str_idx_row]
        curr_cell.value = 'Field value'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        for d_const_instr in d_sw_constr['instr']:

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = d_const_instr['mnem']
            curr_cell.font = Font(name='Consolas', size='11')

            cell_font = Font(name='Calibri', size='11')

            if 'fields' in d_const_instr:

                for str_field_name in d_const_instr['fields']:

                    curr_cell = work_sheet['C' + str_idx_row]
                    curr_cell.value = str_field_name
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['D' + str_idx_row]
                    curr_cell.value = d_const_instr['fields'][str_field_name]
                    curr_cell.font = cell_font

                    i_idx_row += 1
                    str_idx_row = str(i_idx_row)
            else:
                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font

                curr_cell = work_sheet['D' + str_idx_row]
                curr_cell.value = '-'
                curr_cell.font = cell_font

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_instr_off(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_font = Font(name='Calibri', size='11', bold=True)
        cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Included sequence'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

        curr_cell = work_sheet['G' + str_idx_row]
        curr_cell.value = 'Excluded sequences'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':D' + str_idx_row)

        curr_cell = work_sheet['G' + str_idx_row]
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('G' + str_idx_row + ':H' + str_idx_row)

        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Mnemonic 1'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

        curr_cell = work_sheet['C' + str_idx_row]
        curr_cell.value = 'Mnemonic 2'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        curr_cell = work_sheet['D' + str_idx_row]
        curr_cell.value = 'Address offset'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        curr_cell = work_sheet['G' + str_idx_row]
        curr_cell.value = 'Mnemonic 1'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        curr_cell = work_sheet['H' + str_idx_row]
        curr_cell.value = 'Mnemonic 2'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_font = Font(name='Consolas', size='11')

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = d_sw_constr['instr']['incl']['mnem1']
        curr_cell.font = cell_font

        curr_cell = work_sheet['C' + str_idx_row]
        curr_cell.value = d_sw_constr['instr']['incl']['mnem2']
        curr_cell.font = cell_font

        curr_cell = work_sheet['D' + str_idx_row]
        curr_cell.value = "0x%08x" % d_sw_constr['instr']['off']
        curr_cell.font = cell_font

        for str_mnem1 in d_sw_constr['instr']['excl']['mnem1']:
            for str_mnem2 in d_sw_constr['instr']['excl']['mnem2']:
                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = str_mnem1
                curr_cell.font = cell_font

                curr_cell = work_sheet['H' + str_idx_row]
                curr_cell.value = str_mnem2
                curr_cell.font = cell_font

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_comm_str(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_font = Font(name='Calibri', size='11', bold=True)
        cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Strings in sequential order'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':C' + str_idx_row)

        curr_cell = work_sheet['D' + str_idx_row]
        curr_cell.value = 'Exact/Minimum number'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('D' + str_idx_row + ':E' + str_idx_row)

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        for d_comm_str in d_sw_constr['comm']:
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = d_comm_str['string']
            curr_cell.font = Font(name='Consolas', size='11')

            if d_comm_str['min']:
                str_number = 'Minimum'
            else:
                str_number = 'Exact'

            curr_cell = work_sheet['D' + str_idx_row]
            curr_cell.value = str_number + ' number of ' + str(d_comm_str['number'])
            curr_cell.font = Font(name='Calibri', size='11')

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

        i_idx_row += 1

        return i_idx_row

    def _report_verif_constr_log_comp_opt(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        cell_font = Font(name='Calibri', size='11', bold=True)
        cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Source files language'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill
        work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

        curr_cell = work_sheet['C' + str_idx_row]
        curr_cell.value = 'Compilation options'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        curr_cell = work_sheet['D' + str_idx_row]
        curr_cell.value = 'Use/Unuse'
        curr_cell.font = cell_font
        curr_cell.fill = cell_fill

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = d_sw_constr['comp_opt']['lang']
        curr_cell.font = Font(name='Calibri', size='11')

        for str_opt, b_use in d_sw_constr['comp_opt']['opt'].items():

            curr_cell = work_sheet['C' + str_idx_row]
            curr_cell.value = str_opt
            curr_cell.font = Font(name='Consolas', size='11')

            curr_cell = work_sheet['D' + str_idx_row]
            if b_use:
                curr_cell.value = 'Use'
            else:
                curr_cell.value = 'Unuse'
            curr_cell.font = Font(name='Calibri', size='11')

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

        work_sheet.row_dimensions[i_idx_row].height = 4
        i_idx_row += 2

        return i_idx_row

    def _report_verif_constr_log_occur_warn(self, work_sheet, i_idx_row, d_sw_constr):

        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        str_file_names = ', '.join(d_sw_constr['files'])
        curr_cell.value = 'Files ' + str_file_names
        curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1

        return i_idx_row

    def _report_result_constr(self, work_sheet, i_idx_row, str_constr_type, d_sw_constr_result):

        if SwConst.SW_CONST_EXE_MAP_LIM_SECT == str_constr_type:
            self._report_result_constr_map_lim_sect(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_EXE_MAP_OFF_SECT == str_constr_type:
            self._report_result_constr_map_off_sect(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_EXE_MAP_OCCUR_SYMB == str_constr_type:
            self._report_result_constr_map_occur_symb(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_EXE_INSTR_MAND == str_constr_type:
            self._report_result_constr_instr_mand(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_EXE_INSTR_FORB == str_constr_type or SwConst.SW_CONST_EXE_INSTR_OCCUR == str_constr_type:
            self._report_result_constr_instr_forb_occur(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_EXE_INSTR_OFF == str_constr_type:
            self._report_result_constr_instr_off(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_EXE_COMM_STR == str_constr_type:
            self._report_result_constr_comm_str(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_LOG_COMP_OPT == str_constr_type:
            self._report_result_constr_log_comp_opt(work_sheet, i_idx_row, d_sw_constr_result)
        elif SwConst.SW_CONST_LOG_OCCUR_WARN == str_constr_type:
            self._report_result_constr_log_occur_warn(work_sheet, i_idx_row, d_sw_constr_result)
        else:
            pass

    def _report_result_constr_map_lim_sect(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name in d_sw_constr_result['map_lim_sect']:

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Section'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['C' + str_idx_row]
            curr_cell.value = 'Start symbols'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('C' + str_idx_row + ':D' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Start address'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['F' + str_idx_row]
            curr_cell.value = 'End symbols'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('F' + str_idx_row + ':G' + str_idx_row)

            curr_cell = work_sheet['H' + str_idx_row]
            curr_cell.value = 'End address'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Consolas', size='11')
            cell_align = Alignment(horizontal='left', vertical='top')

            for str_sect_name, d_lim_sect in d_sw_constr_result['map_lim_sect'][str_file_name].items():

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = str_sect_name
                curr_cell.font = cell_font
                curr_cell.alignment = cell_align

                if 'start' in d_lim_sect:
                    curr_cell = work_sheet['C' + str_idx_row]
                    curr_cell.value = ', '.join(d_lim_sect['start']['symb'])
                    curr_cell.alignment = Alignment(wrapText=True)
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_lim_sect['start']['addr']
                    curr_cell.font = cell_font
                    curr_cell.alignment = cell_align
                else:
                    curr_cell = work_sheet['C' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font
                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font

                if 'end' in d_lim_sect:
                    curr_cell = work_sheet['F' + str_idx_row]
                    curr_cell.value = ', '.join(d_lim_sect['end']['symb'])
                    curr_cell.alignment = Alignment(vertical='top', wrapText=True)
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['H' + str_idx_row]
                    curr_cell.value = d_lim_sect['end']['addr']
                    curr_cell.font = cell_font
                    curr_cell.alignment = cell_align
                else:
                    curr_cell = work_sheet['F' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font
                    curr_cell = work_sheet['H' + str_idx_row]
                    curr_cell.value = '-'
                    curr_cell.font = cell_font

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7

            i_idx_row += 1

    def _report_result_constr_map_off_sect(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_sect_name, l_symb_names_addr_off in d_sw_constr_result['map_off_sect'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = str_sect_name + ' section'
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Symbols'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':D' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'File'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('E' + str_idx_row + ':F' + str_idx_row)

            curr_cell = work_sheet['G' + str_idx_row]
            curr_cell.value = 'Address offset'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            for d_symb_names_addr_off in l_symb_names_addr_off:

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = ', '.join(d_symb_names_addr_off['symb_names'])
                curr_cell.font = Font(name='Consolas', size='11')
                cell_align = Alignment(horizontal='left', vertical='bottom', wrapText=True)

                for str_file_name, str_symb_addr in d_symb_names_addr_off['symb_addr'].items():

                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = str_file_name
                    curr_cell.font = Font(name='Calibri', size='11')

                    curr_cell = work_sheet['G' + str_idx_row]
                    curr_cell.value = str_symb_addr
                    curr_cell.font = Font(name='Consolas', size='11')

                    i_idx_row += 1
                    str_idx_row = str(i_idx_row)

            i_idx_row += 1

            work_sheet.row_dimensions[i_idx_row].height = 7

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

    def _report_result_constr_map_occur_symb(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name in d_sw_constr_result['map_occur_symb']:

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Section'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['C' + str_idx_row]
            curr_cell.value = 'Symbol'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['D' + str_idx_row]
            curr_cell.value = 'Address'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Size'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Consolas', size='11')

            for str_sect_name, d_map_occur_symb in d_sw_constr_result['map_occur_symb'][str_file_name].items():

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = str_sect_name
                curr_cell.font = cell_font

                for str_symb_name, d_symb_addr_size in d_map_occur_symb.items():

                    curr_cell = work_sheet['C' + str_idx_row]
                    curr_cell.value = str_symb_name
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['D' + str_idx_row]
                    curr_cell.value = d_symb_addr_size['addr']
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_symb_addr_size['size']
                    curr_cell.font = cell_font

                    i_idx_row += 1
                    str_idx_row = str(i_idx_row)

                work_sheet.row_dimensions[i_idx_row].height = 7
                i_idx_row += 1

    def _report_result_constr_instr_mand(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name, d_symb_name_addr_instr in d_sw_constr_result['instr_mand'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            if d_symb_name_addr_instr:

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Symbols'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Instructions'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('G' + str_idx_row + ':J' + str_idx_row)

                work_sheet.row_dimensions[i_idx_row].height = 4

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Name'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':D' + str_idx_row)

                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = 'Address'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Address'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['H' + str_idx_row]
                curr_cell.value = 'Mnemonic'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['I' + str_idx_row]
                curr_cell.value = 'Field name'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['J' + str_idx_row]
                curr_cell.value = 'Field value'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                for str_symb_name in d_symb_name_addr_instr:

                    cell_font = Font(name='Consolas', size='11')

                    curr_cell = work_sheet['A' + str_idx_row]
                    curr_cell.value = str_symb_name
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_symb_name_addr_instr[str_symb_name]['addr']
                    curr_cell.font = cell_font

                    for str_instr_addr, d_instr in d_symb_name_addr_instr[str_symb_name]['instr'].items():

                        cell_font = Font(name='Calibri', size='11')

                        curr_cell = work_sheet['G' + str_idx_row]
                        curr_cell.value = str_instr_addr
                        curr_cell.font = cell_font

                        curr_cell = work_sheet['H' + str_idx_row]
                        curr_cell.value = d_instr['mnem']
                        curr_cell.font = cell_font

                        if 'fields' in d_instr:
                            for str_field_name in d_instr['fields']:
                                curr_cell = work_sheet['I' + str_idx_row]
                                curr_cell.value = str_field_name
                                curr_cell.font = cell_font

                                curr_cell = work_sheet['J' + str_idx_row]
                                curr_cell.value = d_instr['fields'][str_field_name]
                                curr_cell.font = cell_font

                                i_idx_row += 1
                                str_idx_row = str(i_idx_row)
                        else:
                            curr_cell = work_sheet['I' + str_idx_row]
                            curr_cell.value = '-'
                            curr_cell.font = cell_font

                            curr_cell = work_sheet['J' + str_idx_row]
                            curr_cell.value = '-'
                            curr_cell.font = cell_font

                            i_idx_row += 1
                            str_idx_row = str(i_idx_row)

            else:

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'No occurence found'
                curr_cell.font = Font(name='Arial', size='11', italic=True)

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7
            i_idx_row += 1

    def _report_result_constr_instr_forb_occur(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name, d_symb_name_addr_instr in d_sw_constr_result['instr_forb_occur'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            if d_symb_name_addr_instr:

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Symbols'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Instructions'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('G' + str_idx_row + ':J' + str_idx_row)

                work_sheet.row_dimensions[i_idx_row].height = 4

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Name'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':D' + str_idx_row)

                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = 'Address'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Address'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['H' + str_idx_row]
                curr_cell.value = 'Mnemonic'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['I' + str_idx_row]
                curr_cell.value = 'Field name'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['J' + str_idx_row]
                curr_cell.value = 'Field value'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                for str_symb_name in d_symb_name_addr_instr:

                    cell_font = Font(name='Consolas', size='11')

                    curr_cell = work_sheet['A' + str_idx_row]
                    curr_cell.value = str_symb_name
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_symb_name_addr_instr[str_symb_name]['addr']
                    curr_cell.font = cell_font

                    for str_instr_addr, d_instr in d_symb_name_addr_instr[str_symb_name]['instr'].items():

                        cell_font = Font(name='Calibri', size='11')

                        curr_cell = work_sheet['G' + str_idx_row]
                        curr_cell.value = str_instr_addr
                        curr_cell.font = cell_font

                        curr_cell = work_sheet['H' + str_idx_row]
                        curr_cell.value = d_instr['mnem']
                        curr_cell.font = cell_font

                        if 'fields' in d_instr:
                            for str_field_name in d_instr['fields']:
                                curr_cell = work_sheet['I' + str_idx_row]
                                curr_cell.value = str_field_name
                                curr_cell.font = cell_font

                                curr_cell = work_sheet['J' + str_idx_row]
                                curr_cell.value = d_instr['fields'][str_field_name]
                                curr_cell.font = cell_font

                                i_idx_row += 1
                                str_idx_row = str(i_idx_row)
                        else:
                            curr_cell = work_sheet['I' + str_idx_row]
                            curr_cell.value = '-'
                            curr_cell.font = cell_font

                            curr_cell = work_sheet['J' + str_idx_row]
                            curr_cell.value = '-'
                            curr_cell.font = cell_font

                            i_idx_row += 1
                            str_idx_row = str(i_idx_row)

            else:

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'No occurence found'
                curr_cell.font = Font(name='Arial', size='11', italic=True)

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7
            i_idx_row += 1

    def _report_result_constr_instr_off(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name, d_symb_name_instr_addr_off in d_sw_constr_result['instr_off'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            if d_symb_name_instr_addr_off:

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Symbols'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Instructions'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('G' + str_idx_row + ':I' + str_idx_row)

                work_sheet.row_dimensions[i_idx_row].height = 4

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Name'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':D' + str_idx_row)

                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = 'Address'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Address for mnemonic 1'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['H' + str_idx_row]
                curr_cell.value = 'Address for mnemonic 2'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                curr_cell = work_sheet['I' + str_idx_row]
                curr_cell.value = 'Address offset'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                for str_symb_name in d_symb_name_instr_addr_off:

                    cell_font = Font(name='Consolas', size='11')

                    curr_cell = work_sheet['A' + str_idx_row]
                    curr_cell.value = str_symb_name
                    curr_cell.font = cell_font

                    curr_cell = work_sheet['E' + str_idx_row]
                    curr_cell.value = d_symb_name_instr_addr_off[str_symb_name]['symb_addr']
                    curr_cell.font = cell_font

                    for d_instr_addr_off in d_symb_name_instr_addr_off[str_symb_name]['inst_addr']:

                        cell_font = Font(name='Calibri', size='11')

                        curr_cell = work_sheet['G' + str_idx_row]
                        curr_cell.value = d_instr_addr_off['mnem1']
                        curr_cell.font = cell_font

                        curr_cell = work_sheet['H' + str_idx_row]
                        curr_cell.value = d_instr_addr_off['mnem2']
                        curr_cell.font = cell_font

                        curr_cell = work_sheet['I' + str_idx_row]
                        curr_cell.value = d_instr_addr_off['off']
                        curr_cell.font = cell_font

                        i_idx_row += 1
                        str_idx_row = str(i_idx_row)

            else:

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'No occurence found'
                curr_cell.font = Font(name='Arial', size='11', italic=True)

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7
            i_idx_row += 1

    def _report_result_constr_comm_str(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name, d_comm_str_list in d_sw_constr_result['comm_str'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Strings in sequential order'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':D' + str_idx_row)

            curr_cell = work_sheet['E' + str_idx_row]
            curr_cell.value = 'Number'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            for d_comm_str in d_comm_str_list:
                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = d_comm_str['string']
                curr_cell.font = Font(name='Consolas', size='11')

                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = d_comm_str['number']
                curr_cell.font = Font(name='Calibri', size='11')

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7
            i_idx_row += 1

    def _report_result_constr_log_comp_opt(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name, d_lang_comp_opt  in d_sw_constr_result['log_comp_opt'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            cell_font = Font(name='Calibri', size='11', bold=True)
            cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'Source files language'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('A' + str_idx_row + ':B' + str_idx_row)

            curr_cell = work_sheet['C' + str_idx_row]
            curr_cell.value = 'Compilation options'
            curr_cell.font = cell_font
            curr_cell.fill = cell_fill
            work_sheet.merge_cells('C' + str_idx_row + ':E' + str_idx_row)

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = d_lang_comp_opt['lang']
            curr_cell.font = Font(name='Calibri', size='11')

            for str_opt, l_opt_text in d_lang_comp_opt['comp_opt'].items():

                curr_cell = work_sheet['C' + str_idx_row]
                curr_cell.value = str_opt
                curr_cell.font = curr_cell.font = Font(name='Consolas', size='11')

                if l_opt_text:
                    for str_opt_text in l_opt_text:

                        curr_cell = work_sheet['D' + str_idx_row]
                        curr_cell.value = str_opt_text
                        curr_cell.font = Font(name='Calibri', size='11')
                        work_sheet.merge_cells('D' + str_idx_row + ':E' + str_idx_row)
                        curr_cell.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
                        curr_cell.style = 'Normal'

                        i_idx_row += 1
                        str_idx_row = str(i_idx_row)
                else:
                    i_idx_row += 1
                    str_idx_row = str(i_idx_row)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

    def _report_result_constr_log_occur_warn(self, work_sheet, i_idx_row, d_sw_constr_result):

        for str_file_name, d_occur_warn in d_sw_constr_result['log_occur_warn'].items():

            str_idx_row = str(i_idx_row)
            curr_cell = work_sheet['A' + str_idx_row]
            curr_cell.value = 'File ' + str_file_name
            curr_cell.font = Font(name='Helvetica Narrow', size='11', bold=True)

            i_idx_row += 1
            work_sheet.row_dimensions[i_idx_row].height = 4

            i_idx_row += 1
            str_idx_row = str(i_idx_row)

            if d_occur_warn:

                cell_font = Font(name='Calibri', size='11', bold=True)
                cell_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                cell_alignment = Alignment(horizontal='right')

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'Source file'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill
                work_sheet.merge_cells('A' + str_idx_row + ':C' + str_idx_row)

                curr_cell = work_sheet['D' + str_idx_row]
                curr_cell.value = 'Line'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill
                curr_cell.alignment = cell_alignment

                curr_cell = work_sheet['E' + str_idx_row]
                curr_cell.value = 'Column'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill
                curr_cell.alignment = cell_alignment

                curr_cell = work_sheet['G' + str_idx_row]
                curr_cell.value = 'Text'
                curr_cell.font = cell_font
                curr_cell.fill = cell_fill

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

                for str_src_file_name in d_occur_warn:

                    curr_cell = work_sheet['A' + str_idx_row]
                    curr_cell.value = str_src_file_name
                    curr_cell.font = Font(name='Consolas', size='11')

                    cell_font = Font(name='Calibri', size='11')

                    for str_line_number in d_occur_warn[str_src_file_name]:
                        i_line_number = int(str_line_number)

                        for str_col_number in d_occur_warn[str_src_file_name][str_line_number]:
                            if '' == str_col_number:
                                for str_text in d_occur_warn[str_src_file_name][str_line_number][str_col_number]:

                                    curr_cell = work_sheet['D' + str_idx_row]
                                    curr_cell.value = i_line_number
                                    curr_cell.font = cell_font
                                    curr_cell.alignment = cell_alignment

                                    curr_cell = work_sheet['E' + str_idx_row]
                                    curr_cell.value = '-'
                                    curr_cell.font = cell_font
                                    curr_cell.alignment = cell_alignment

                                    curr_cell = work_sheet['G' + str_idx_row]
                                    curr_cell.value = str_text
                                    curr_cell.font = cell_font
                                    curr_cell.alignment = Alignment(horizontal='left', vertical='bottom', wrapText=True)

                                    i_idx_row += 1
                                    str_idx_row = str(i_idx_row)

                            else:

                                curr_cell = work_sheet['D' + str_idx_row]
                                curr_cell.value = i_line_number
                                curr_cell.font = cell_font
                                curr_cell.alignment = cell_alignment

                                i_col_number = int(str_col_number)

                                curr_cell = work_sheet['E' + str_idx_row]
                                curr_cell.value = i_col_number
                                curr_cell.font = cell_font
                                curr_cell.alignment = cell_alignment

                                str_text = d_occur_warn[str_src_file_name][str_line_number][str_col_number]

                                curr_cell = work_sheet['G' + str_idx_row]
                                curr_cell.value = str_text
                                curr_cell.font = cell_font
                                curr_cell.alignment = Alignment(horizontal='left', vertical='bottom', wrapText=True)

                                i_idx_row += 1
                                str_idx_row = str(i_idx_row)

            else:

                curr_cell = work_sheet['A' + str_idx_row]
                curr_cell.value = 'No warning found'
                curr_cell.font = Font(name='Arial', size='11', italic=True)

                i_idx_row += 1
                str_idx_row = str(i_idx_row)

            work_sheet.row_dimensions[i_idx_row].height = 7
            i_idx_row += 1

    def _add_constraint_id(self, work_sheet, str_constr_id):
        curr_cell = work_sheet['A1']
        curr_cell.value = 'Constraint ' + str_constr_id
        curr_cell.font = Font(name='Calibri', size='16', bold=True, color='244062')

    def _add_constraint_type(self, work_sheet, str_constr_type_text):

        curr_cell = work_sheet['A2']
        curr_cell.value = str_constr_type_text
        curr_cell.font = Font(name='Calibri', size='14', color='366092')

    def _add_verif_banner(self, work_sheet, str_constr_type):

        curr_cell = work_sheet['A4']
        curr_cell.value = 'Verification'
        curr_cell.font = Font(name='Segoe UI Semilight', size='11', color='FFFFFF', bold=True)
        curr_cell.fill = PatternFill(start_color='244062', end_color='244062', fill_type='solid')
        work_sheet.merge_cells('A4:B4')

        curr_cell = work_sheet['A5']
        curr_cell.fill = PatternFill(start_color='244062', end_color='244062', fill_type='solid')

        if SwConst.SW_CONST_EXE_MAP_LIM_SECT == str_constr_type:
            work_sheet.merge_cells('A5:H5')
        elif SwConst.SW_CONST_EXE_MAP_OFF_SECT == str_constr_type:
            work_sheet.merge_cells('A5:G5')
        elif SwConst.SW_CONST_EXE_MAP_OCCUR_SYMB == str_constr_type:
            work_sheet.merge_cells('A5:E5')
        elif SwConst.SW_CONST_EXE_INSTR_MAND == str_constr_type or SwConst.SW_CONST_EXE_INSTR_FORB == str_constr_type \
                or SwConst.SW_CONST_EXE_INSTR_OCCUR == str_constr_type:
            work_sheet.merge_cells('A5:J5')
        elif SwConst.SW_CONST_EXE_INSTR_OFF == str_constr_type:
            work_sheet.merge_cells('A5:I5')
        elif SwConst.SW_CONST_EXE_COMM_STR == str_constr_type:
            work_sheet.merge_cells('A5:E5')
        elif SwConst.SW_CONST_LOG_COMP_OPT == str_constr_type:
            work_sheet.merge_cells('A5:E5')
        elif SwConst.SW_CONST_LOG_OCCUR_WARN == str_constr_type:
            work_sheet.merge_cells('A5:G5')
        else:
            pass

        work_sheet.row_dimensions[5].height = 4
        work_sheet.row_dimensions[6].height = 4

        return 7

    def _add_result_banner(self, work_sheet, i_idx_row, e_result, str_constr_type):

        str_idx_row = str(i_idx_row)
        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.value = 'Result'
        curr_cell.font = Font(name='Segoe UI Semilight', size='11', color='FFFFFF', bold=True)
        curr_cell.fill = PatternFill(start_color='403151', end_color='403151', fill_type='solid')

        curr_cell = work_sheet['B' + str_idx_row]
        curr_cell.font = Font(name='Arial', size='11', color='000000', bold=True)
        curr_cell.alignment = Alignment(horizontal='center')

        curr_cell.value = e_result.value

        if e_result == SwConstResult.SW_CONST_RESULT_OK:
            curr_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        elif e_result == SwConstResult.SW_CONST_RESULT_NOK:
            curr_cell.fill = PatternFill(start_color='FF3300', end_color='FF3300', fill_type='solid')
        elif e_result == SwConstResult.SW_CONST_RESULT_TBA:
            curr_cell.fill = PatternFill(start_color='FFDD79', end_color='FFDD79', fill_type='solid')
        else:
            pass

        i_idx_row += 1
        str_idx_row = str(i_idx_row)

        curr_cell = work_sheet['A' + str_idx_row]
        curr_cell.fill = PatternFill(start_color='403151', end_color='403151', fill_type='solid')

        if SwConst.SW_CONST_EXE_MAP_LIM_SECT == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':H' + str_idx_row)
        elif SwConst.SW_CONST_EXE_MAP_OFF_SECT == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':G' + str_idx_row)
        elif SwConst.SW_CONST_EXE_MAP_OCCUR_SYMB == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)
        elif SwConst.SW_CONST_EXE_INSTR_MAND == str_constr_type or SwConst.SW_CONST_EXE_INSTR_FORB == str_constr_type \
                or SwConst.SW_CONST_EXE_INSTR_OCCUR == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':J' + str_idx_row)
        elif SwConst.SW_CONST_EXE_INSTR_OFF == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':I' + str_idx_row)
        elif SwConst.SW_CONST_EXE_COMM_STR == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)
        elif SwConst.SW_CONST_LOG_COMP_OPT == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':E' + str_idx_row)
        elif SwConst.SW_CONST_LOG_OCCUR_WARN == str_constr_type:
            work_sheet.merge_cells('A' + str_idx_row + ':G' + str_idx_row)
        else:
            pass

        work_sheet.row_dimensions[i_idx_row].height = 4
        i_idx_row += 1
        work_sheet.row_dimensions[i_idx_row].height = 4

        i_idx_row += 1
        return i_idx_row

    def _set_page_format(self, work_sheet, str_constr_type):

        work_sheet.set_printer_settings(Worksheet.PAPERSIZE_A4, Worksheet.ORIENTATION_LANDSCAPE)
        work_sheet.page_margins.left = 0.6
        work_sheet.page_margins.right = 0.6
        work_sheet.page_margins.top = 0.7
        work_sheet.sheet_view.view = 'pageLayout'
        work_sheet.sheet_view.showGridLines = False

        work_sheet.column_dimensions['A'].width = 14
        work_sheet.column_dimensions['B'].width = 6

        if SwConst.SW_CONST_EXE_MAP_LIM_SECT == str_constr_type:

            work_sheet.column_dimensions['C'].width = 37
            work_sheet.column_dimensions['D'].width = 1
            work_sheet.column_dimensions['E'].width = 20
            work_sheet.column_dimensions['F'].width = 37
            work_sheet.column_dimensions['G'].width = 1
            work_sheet.column_dimensions['H'].width = 17

        elif SwConst.SW_CONST_EXE_MAP_OCCUR_SYMB == str_constr_type:

            work_sheet.column_dimensions['C'].width = 57
            work_sheet.column_dimensions['D'].width = 14
            work_sheet.column_dimensions['E'].width = 14

        elif SwConst.SW_CONST_EXE_MAP_OFF_SECT == str_constr_type:

            work_sheet.column_dimensions['C'].width = 19
            work_sheet.column_dimensions['D'].width = 39
            work_sheet.column_dimensions['E'].width = 14
            work_sheet.column_dimensions['F'].width = 25
            work_sheet.column_dimensions['G'].width = 14

        elif SwConst.SW_CONST_EXE_INSTR_MAND == str_constr_type or SwConst.SW_CONST_EXE_INSTR_FORB == str_constr_type \
                or SwConst.SW_CONST_EXE_INSTR_OCCUR == str_constr_type:

            work_sheet.column_dimensions['C'].width = 16
            work_sheet.column_dimensions['D'].width = 16
            work_sheet.column_dimensions['E'].width = 14
            work_sheet.column_dimensions['F'].width = 1
            work_sheet.column_dimensions['G'].width = 14
            work_sheet.column_dimensions['H'].width = 20
            work_sheet.column_dimensions['I'].width = 16
            work_sheet.column_dimensions['J'].width = 16

        elif SwConst.SW_CONST_EXE_INSTR_OFF == str_constr_type:

            work_sheet.column_dimensions['C'].width = 16
            work_sheet.column_dimensions['D'].width = 16
            work_sheet.column_dimensions['E'].width = 14
            work_sheet.column_dimensions['F'].width = 1
            work_sheet.column_dimensions['G'].width = 21
            work_sheet.column_dimensions['H'].width = 21
            work_sheet.column_dimensions['I'].width = 21

        elif SwConst.SW_CONST_EXE_COMM_STR == str_constr_type:

            work_sheet.column_dimensions['C'].width = 89
            work_sheet.column_dimensions['D'].width = 12
            work_sheet.column_dimensions['E'].width = 9

        elif SwConst.SW_CONST_LOG_COMP_OPT == str_constr_type:

            work_sheet.column_dimensions['C'].width = 38
            work_sheet.column_dimensions['D'].width = 11
            work_sheet.column_dimensions['E'].width = 64

        elif SwConst.SW_CONST_LOG_OCCUR_WARN == str_constr_type:

            work_sheet.column_dimensions['C'].width = 22
            work_sheet.column_dimensions['D'].width = 9
            work_sheet.column_dimensions['E'].width = 9
            work_sheet.column_dimensions['F'].width = 1
            work_sheet.column_dimensions['G'].width = 72

    def _save(self):

        if 'Sheet' in self.work_book:
            ws_sheet = self.work_book['Sheet']
            self.work_book.remove(ws_sheet)

        str_file_path = self.str_dir_path + '/' + self.str_sys_sw_sw_ver
        str_date_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        str_file_path += '_' + str_date_time + '.xlsx'

        self.work_book.save(str_file_path)
        print(f'\nReport created in directory {f"{os.path.sep}".join(os.path.split(str_file_path)[:-1])}\n'
              f'\twith name: {os.path.split(str_file_path)[-1]}.')
